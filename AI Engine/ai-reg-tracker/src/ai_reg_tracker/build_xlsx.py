"""Render dist/regulatory_map.xlsx - the corrected Regulatory Map and a
Regulatory Horizon sheet, generated from the records (never hand-kept)."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .loader import load_all

REPO_ROOT = Path(__file__).resolve().parents[2]
DIST = REPO_ROOT / "dist"

NAVY = "1F3A5F"
ROW = "EAF0F6"
LINE = "BBC7CB"
FNT = "Calibri"

hdr_font = Font(name=FNT, size=11, bold=True, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor=NAVY)
base = Font(name=FNT, size=11)
bold = Font(name=FNT, size=11, bold=True)
side = Side(style="thin", color=LINE)
border = Border(left=side, right=side, top=side, bottom=side)
wrap = Alignment(wrap_text=True, vertical="top")
ctr = Alignment(horizontal="center", vertical="center")


def _sheet(wb, title, headers, widths):
    ws = wb.create_sheet(title)
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.border = border
        c.alignment = ctr
        ws.column_dimensions[get_column_letter(j)].width = widths[j - 1]
    ws.freeze_panes = "A2"
    return ws


def _write(ws, rows):
    for i, row in enumerate(rows, start=2):
        for j, v in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=v)
            c.font = bold if j == 1 else base
            c.border = border
            c.alignment = wrap
        if i % 2 == 0:
            for j in range(1, len(row) + 1):
                if not ws.cell(row=i, column=j).fill.fgColor.rgb or \
                        ws.cell(row=i, column=j).fill.patternType is None:
                    ws.cell(row=i, column=j).fill = PatternFill("solid", fgColor=ROW)


def main() -> int:
    regs = load_all()
    wb = Workbook()
    wb.remove(wb.active)

    # --- Regulatory Map ---
    rmap = _sheet(
        wb, "Regulatory Map",
        ["Regulation", "Body", "Jurisdiction", "Regime", "Status",
         "Field groups", "Domain", "Summary", "Source"],
        [30, 24, 12, 14, 13, 12, 22, 60, 30],
    )
    map_rows = []
    for r in sorted(regs, key=lambda x: (x.jurisdiction.value, x.name)):
        map_rows.append([
            r.name, r.body, r.jurisdiction.value, r.regime_type.value, r.status.value,
            ", ".join(r.field_groups) or "-",
            ", ".join(d.value for d in r.domain),
            r.summary,
            str(r.source_urls[0]),
        ])
    _write(rmap, map_rows)

    # --- Regulatory Horizon (records with dated deadlines or recent changes) ---
    rh = _sheet(
        wb, "Regulatory Horizon",
        ["Date", "Regulation", "Body", "Event", "Field groups"],
        [13, 30, 24, 50, 14],
    )
    horizon = []
    for r in regs:
        for d in r.dates.deadlines:
            horizon.append([str(d.date), r.name, r.body, d.note, ", ".join(r.field_groups) or "-"])
        for c in r.changelog:
            horizon.append([str(c.date), r.name, r.body, c.change, ", ".join(r.field_groups) or "-"])
    horizon.sort(key=lambda x: x[0], reverse=True)
    _write(rh, horizon)

    DIST.mkdir(exist_ok=True)
    out = DIST / "regulatory_map.xlsx"
    wb.save(out)
    print(f"wrote {out} ({len(map_rows)} regs, {len(horizon)} horizon rows)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
